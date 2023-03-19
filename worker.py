import openpyxl as pyxl
from time import time,sleep
from os import path, getenv,getcwd, mkdir, rename
from random import uniform
from io import BytesIO
from constants import *

def create_bins():
	for i in range(0, DML_STAGING_LOCATION.split(APP_URI)[1].count("\\")):
		mkdir(path.join(APP_URI,"\\".join(DML_STAGING_LOCATION.split(APP_URI+"\\")[1].split("\\")[0:i+1]))) if not(path.isdir(path.join(APP_URI,"\\".join(DML_STAGING_LOCATION.split(APP_URI+"\\")[1].split("\\")[0:i+1])))) else next

	for i in range(0, PROVIDERS["MPS"]["raw_dump"].split(APP_URI)[1].count("\\")):
		mkdir(path.join(APP_URI,"\\".join(PROVIDERS["MPS"]["raw_dump"].split(APP_URI+"\\")[1].split("\\")[0:i+1]))) if not(path.isdir(path.join(APP_URI,"\\".join(PROVIDERS["MPS"]["raw_dump"].split(APP_URI+"\\")[1].split("\\")[0:i+1])))) else next

	mkdir(PROVIDERS["MPS"]["csv_dump"]) if not(path.isdir(PROVIDERS["MPS"]["csv_dump"])) else next

def parse_mps_file():

	landed_file_uri = path.join(FILE_LANDING_URI,PROVIDERS["MPS"]["filename"])

	sql_tablename = "TRANSACTIONS_DUMP"
	header = "FLOW_ID, EXEC, PROVIDER, TR_DATE, TR_WEEK, TR_MONTH, TR_DAY, TR_WEEKDAY, TR_HOUR, TR_TYPE, TR_LOCATION, TR_BENEFICIARY, TR_METHOD, TR_INFO, TR_VALUE, TR_UOM"
	revenues_owner = "SELF"

	try:
		with open(landed_file_uri, "rb") as f:
			wb = pyxl.load_workbook(BytesIO(f.read()), data_only=True)
	except FileNotFoundError:
		raise FileNotFoundError
	except Exception:
		sleep(5) # give time to complete dowloading the file in case of concurrent file dowload completion and pyxl load_workbook
		try:
			with open(landed_file_uri, "rb") as f:
				wb = pyxl.load_workbook(BytesIO(f.read()), data_only=True)
		except Exception as e:
			raise e # raise unknown exception and halt the caller process
		
	else:
		print("1 new transactions file found in Download folder..")
		data = wb[wb.sheetnames[0]]

		exec = int(time())

		processed_filename = str(exec)+"_"+landed_file_uri.split("\\")[-1].replace(" ","_")

		sql_insert_name = f"INSERT_{sql_tablename}_{exec}.sql"
		sql_insert_dml = "" # empty insert sql statement into target tablename

		csv_name = f"TRANSACTIONS_{exec}.csv"
		csv = header + "\n" # initialise empty csv with specified header

		print("parsing new report ..")
		for r in data.iter_rows(min_col=3, max_col=7, min_row=20, values_only=True):
			
			r = ["TRANSACTIONS", exec, PROVIDERS["MPS"]["name"]] + list(r) + ["EUR"] # adds metadata to  payload making it unique even if extraction is repeated w/ same parameters
			
			if "Totale" in r[5]:
				wb.close()
				break

			r[5] = r[5].replace("\n"," ") # removes \n inserted by the provider when transaction description gets verbose
			r.pop(6) # remove empty cell for each row, provider-specific file structure

			r.insert(4,int(r[3].strftime("%W"))) # weeknum
			r.insert(5,int(r[3].strftime("%m"))) # monthnum
			r.insert(6,int(r[3].strftime("%d"))) # daynum
			r.insert(7,int(r[3].strftime("%w"))) # weekday
			r[3]=r[3].strftime("%Y-%m-%d")

			if "Prelievo " in r[9]:
				r.insert(8,int(r[9].split("Prelievo Self Service Atm ")[1].split(" ")[2].split(":")[0])) # hour
			elif "Bon." in r[9] or  "Storno" in r[9] or "bollo" in r[8] or "Addebito" in r[8]:
				r.insert(8,int(uniform(9, 20))) # hour randomly assigned due to lack of data in provider payload
			else:
				try:
					r.insert(8,int(r[9].split(" Ora ")[1][0:2])) # hour
				except IndexError:
					r.insert(8,None) # hour

			r.insert(9, "expense" if r[-2]<=0 else "revenue") # type

			if "Prelievo " in r[11]:
				r.insert(10," ".join(r[11].split("Prelievo Self Service ")[1].split(" ")[0:2])) # location
			elif "Bon." in r[11]:
				r.insert(10,"Home Banking") # location
			else:
				try:
					r.insert(10,r[11].split(" Loc.")[1].split(" Esercente" if " Esercente" in r[11] else " Imp.")[0].title()) # location
				except IndexError:
					r.insert(10,None) # location

			if r[-2]>0:
				r.insert(11,revenues_owner) # beneficiary
			elif "Prel." in r[11]:
				r.insert(11,"Unknown") # beneficiary
			elif "Bon." in r[11] or "Addebito" in r[11]:
				r.insert(11,r[12].split(" A Favore ")[1].split(" Iban" if " Iban" in r[12] else "Codice")[0].strip(" ")) # beneficiary
			else:
				try:
					r.insert(11,r[12].split("Esercente")[1].split(" Imp.")[0].strip().lstrip(":").lstrip().title()) # beneficiary
				except IndexError:
					r.insert(11,None) # beneficiary

			csv += str(r)[1:-1].replace(" '",' "').replace("',",'",') + "\n"
			sql_insert_dml += f"INSERT INTO {SQL_DB_NAME}.{sql_tablename} ({header}) VALUES ({str(r)[1:-1]});" + "\n"

		try:
			with open(PROVIDERS["MPS"]["csv_dump"]+"\\"+csv_name, "w") as f:
				f.write(csv.replace("\n'",'\n"').replace("'\n",'"\n'))

			with open(DML_STAGING_LOCATION+"\\"+sql_insert_name, "w") as f:
				f.write(sql_insert_dml.replace("None","NULL"))

			rename(src=landed_file_uri,dst=path.join(PROVIDERS["MPS"]["raw_dump"],processed_filename))
		except Exception:
			create_bins()

			with open(PROVIDERS["MPS"]["csv_dump"]+"\\"+csv_name, "w") as f:
				f.write(csv.replace("\n'",'\n"').replace("'\n",'"\n'))

			with open(DML_STAGING_LOCATION+"\\"+sql_insert_name, "w") as f:
				f.write(sql_insert_dml.replace("None","NULL"))

			rename(src=landed_file_uri,dst=path.join(PROVIDERS["MPS"]["raw_dump"],processed_filename))
		finally:
			print("1 new transaction report ready for ingestion!")

